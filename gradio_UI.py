import gradio as gr
import requests
import json
import csv
import openpyxl
# import spacy
import re

# nlp = spacy.load("en_core_web_sm") TODO not used, commented out

# Ollama Docker server details 
# TODO change this and all other variable will follow
OLLAMA_SERVER_URL = "http://127.0.0.1:11434"  # "http://172.17.0.2:11434" for docker network TODO change to match actual ollama/ ollama container IP
MODEL = "llama3.2"

#----- RAG call Here ----------------------------------------------------------------------
import rag
rag = rag.Rag(db="faiss_local",OLLAMA_SERVER_URL=OLLAMA_SERVER_URL,model=MODEL)

#----- Parsing Output Here ----------------------------------------------------------------------
def parse_ollama_response(response):
    """
    Parses the streaming response from Ollama into a clean format.
    """
    try:
        full_response = ""
        response_lines = response.text.split('\n')
        
        for line in response_lines:
            if not line:
                continue
            try:
                data = json.loads(line)
                if 'response' in data:
                    full_response += data['response']
            except json.JSONDecodeError:
                continue
                
        return full_response.strip()
    except Exception as e:
        return f"Error parsing response: {e}"

#----- Input Query Here -------------------------------------------------------------------------
def query_ollama(user_query):
    """
    Process input and query Ollama model.
    """
    #parse user input, clean problematic chars 
    user_query = user_query
    prompt = rag.query_parser(user_query)
    
    #retrieve relevant docs via RAG
    try:
        retrieved_files = rag.submit_query(prompt) #currently just searching by prompt TODO test different combination of keywords
        prompt = f"explain how an attack can happen where: \n{prompt}\n by refering to: \n{retrieved_files}" #TODO change format of prompt, currently used to test on untrained llm
     
    except Exception as error:
            return "error when retrieving relavent docs:\n"+str(error)
    
    # Query Ollama
    url = f"{OLLAMA_SERVER_URL}/api/generate"
    payload = {
        "model": "llama3.2",#TODO change to actual model name
        "prompt": prompt,
        "stream": True
    }
    
    try:
        response = requests.post(url, json=payload, timeout=60)
        response.raise_for_status()
        # Parse response
        cleaned_response = parse_ollama_response(response)
        return cleaned_response if cleaned_response else "No response generated"
        
    except requests.exceptions.RequestException as e:
        return f"Error communicating with Ollama: {str(e)}"
        
#----- Read uploaded CSV file for instructions --------------------------------------------------
def parse_instruction_csv(f):
    query=[]
    result=''
    with open(f,'r') as f1:
        reader=csv.reader(f1)
        for box in reader:
            query.append(box[0].replace('\n','').split('>')) # Cleaning unwanted string and separate by stages 
    print(query)  # Debug: Log parsed instructions
    for line in query:
        for i in line:
            result += "i: "+query_ollama(i) + "\n"   # Send query to Ollama
    return result

# Receive uploaded file and send to llama
def upload_file(f):
    f_parsed = parse_instruction_csv(f)
    return f_parsed

#----- Export editable output to Excel ----------------------------------------------------------
def editable_export_excel(editable_output):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        # FOR OUTPUT IN NEXT LINES
        # lines = editable_output.split('\n')
        # for i, line in enumerate(lines):
        #     ws.cell(row=i+1, column=1).value = line

        ws.cell(row=1, column=1).value = editable_output

        file_name = "output.xlsx"
        wb.save(file_name)
        return gr.File(file_name)
    except Exception as e:
        return str(e)
        
#----- User Interface ---------------------------------------------------------------------------
with gr.Blocks(theme=gr.themes.Citrus()) as demo:
    gr.Markdown("<center><h1>CRISE - Cyber RIsk Scenario Engine</h1></center>")
    gr.Markdown("<center>This is a LLM semi-automation used for generating risk scenarios. Outputs are color-coded for easy reference.<center>")

    with gr.Row():
        with gr.Column():
            input = gr.Textbox(interactive=True, label="Enter Attack Vector", placeholder="Enter text following the format shown below or upload file")
            with gr.Row():
                clear_button = gr.Button("Clear")
                submit_button = gr.Button("Submit")
            gr.Markdown("<b>How to chat with CRISE:</b>")
            gr.Markdown("threat {Threat_Event}<br>cause {Vulnerability}<br>asset {Asset at risk}<br>consequence {example consequences}")
        with gr.Column():
            output = gr.Textbox(label="Editable Output", interactive=True)
            export_button = gr.Button("Export Output - Excel")
            export_file = gr.File(label="Edited Excel File")
            export_button.click(editable_export_excel, inputs=output, outputs=export_file)
            #db create, rag_changes
            # db_create_button = gr.UploadButton("create a new database -Warning: will overwrite current db", type='filepath', file_count="single")
            # db_create_button.upload(fn=rag.create_vector_store, inputs=db_create_button)
            #db update, rag_changes
            with gr.Row():
                db_add_button = gr.UploadButton("add more data to current database", type='filepath', file_count="single")
                db_add_button.upload(fn=rag.add_vector_store, inputs=db_add_button)
                db_get_button = gr.DownloadButton("download currently stored data",value=".\\data.csv")

        def clear_input():
            return ""


        clear_button.click(clear_input, inputs=None, outputs=input)
        submit_button.click(query_ollama, inputs=input, outputs=output, show_progress=True)

#----- Launch the app ---------------------------------------------------------------------------
demo.launch(server_name="0.0.0.0", server_port=7860)

#torch.cuda.empty_cache()  # Add after training completion TODO im not sure what this does so im leaving this here